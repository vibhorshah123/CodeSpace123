"""
Excel Generator Module
Creates Excel reports from comparison results
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List
from datetime import datetime


class ExcelGenerator:
    """Generates Excel reports from comparison data"""
    
    def __init__(self):
        """Initialize Excel generator"""
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.info_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    
    def _apply_header_style(self, ws, row_num: int, max_col: int):
        """Apply header styling to a row"""
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def generate_schema_comparison_report(self, comparison_result: Dict[str, Any], table_name: str, output_file: str, source_env_name: str = None, target_env_name: str = None):
        """
        Generate Excel report for schema comparison
        
        Args:
            comparison_result: Dictionary containing comparison results
            table_name: Name of the table being compared
            output_file: Output Excel file path
            source_env_name: Name of source environment (optional)
            target_env_name: Name of target environment (optional)
        """
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Create Summary sheet
        self._create_summary_sheet(wb, comparison_result, table_name, source_env_name, target_env_name)
        
        # Create Fields Only in Source sheet
        if comparison_result['only_in_source']:
            self._create_only_in_source_sheet(wb, comparison_result, source_env_name)
        
        # Create Fields Only in Target sheet
        if comparison_result['only_in_target']:
            self._create_only_in_target_sheet(wb, comparison_result, target_env_name)
        
        # Create Field Differences sheet
        if comparison_result['field_differences']:
            self._create_field_differences_sheet(wb, comparison_result)
        
        # Create Matching Fields sheet
        if comparison_result['matching_fields']:
            self._create_matching_fields_sheet(wb, comparison_result)
        
        # Save workbook
        wb.save(output_file)
        print(f"  Excel report saved to: {output_file}")
    
    def _create_summary_sheet(self, wb, comparison_result: Dict[str, Any], table_name: str, source_env_name: str = None, target_env_name: str = None):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "Schema Comparison Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:B1')
        
        # Metadata
        row = 3
        ws[f'A{row}'] = "Table Name:"
        ws[f'B{row}'] = table_name
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Source Environment:"
        ws[f'B{row}'] = comparison_result['source_url']
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Target Environment:"
        ws[f'B{row}'] = comparison_result['target_url']
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Report Generated:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws[f'A{row}'].font = Font(bold=True)
        
        # Summary statistics
        row += 2
        ws[f'A{row}'] = "COMPARISON SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Count"
        self._apply_header_style(ws, row, 2)
        
        # Statistics rows
        source_label = source_env_name if source_env_name else "Source"
        target_label = target_env_name if target_env_name else "Target"
        stats = [
            (f"Fields Only in {source_label}", len(comparison_result['only_in_source']), self.warning_fill),
            (f"Fields Only in {target_label}", len(comparison_result['only_in_target']), self.warning_fill),
            ("Fields with Differences", len(comparison_result['field_differences']), self.info_fill),
            ("Matching Fields", len(comparison_result['matching_fields']), self.success_fill)
        ]
        
        for label, value, fill in stats:
            row += 1
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            ws[f'B{row}'].fill = fill
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
        
        self._auto_adjust_columns(ws)
    
    def _create_only_in_source_sheet(self, wb, comparison_result: Dict[str, Any], source_env_name: str = None):
        """Create sheet for fields only in source"""
        sheet_name = f"Only in {source_env_name}" if source_env_name else "Only in Source"
        ws = wb.create_sheet(sheet_name)
        
        # Headers
        ws['A1'] = "Field Logical Name"
        ws['A1'].fill = self.warning_fill
        ws['A1'].font = Font(bold=True)
        ws['A1'].border = self.border
        
        # Data
        for idx, field_name in enumerate(comparison_result['only_in_source'], start=2):
            ws[f'A{idx}'] = field_name
            ws[f'A{idx}'].border = self.border
        
        self._auto_adjust_columns(ws)
    
    def _create_only_in_target_sheet(self, wb, comparison_result: Dict[str, Any], target_env_name: str = None):
        """Create sheet for fields only in target"""
        sheet_name = f"Only in {target_env_name}" if target_env_name else "Only in Target"
        ws = wb.create_sheet(sheet_name)
        
        # Headers
        ws['A1'] = "Field Logical Name"
        ws['A1'].fill = self.warning_fill
        ws['A1'].font = Font(bold=True)
        ws['A1'].border = self.border
        
        # Data
        for idx, field_name in enumerate(comparison_result['only_in_target'], start=2):
            ws[f'A{idx}'] = field_name
            ws[f'A{idx}'].border = self.border
        
        self._auto_adjust_columns(ws)
    
    def _create_field_differences_sheet(self, wb, comparison_result: Dict[str, Any]):
        """Create sheet for field differences"""
        ws = wb.create_sheet("Field Differences")
        
        # Headers
        headers = ["Field Name", "Property", "Source Value", "Target Value"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
        
        self._apply_header_style(ws, 1, len(headers))
        
        # Data
        row = 2
        for field_diff in comparison_result['field_differences']:
            field_name = field_diff['field_name']
            differences = field_diff['differences']
            
            # Write each difference as a row
            for prop_name, values in differences.items():
                ws[f'A{row}'] = field_name
                ws[f'B{row}'] = prop_name
                ws[f'C{row}'] = str(values['source'])
                ws[f'D{row}'] = str(values['target'])
                
                # Apply borders
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = self.border
                
                # Highlight differences
                ws[f'C{row}'].fill = self.info_fill
                ws[f'D{row}'].fill = self.info_fill
                
                row += 1
        
        self._auto_adjust_columns(ws)
    
    def _create_matching_fields_sheet(self, wb, comparison_result: Dict[str, Any]):
        """Create sheet for matching fields"""
        ws = wb.create_sheet("Matching Fields")
        
        # Headers
        ws['A1'] = "Field Logical Name"
        ws['A1'].fill = self.success_fill
        ws['A1'].font = Font(bold=True)
        ws['A1'].border = self.border
        
        # Data
        for idx, field_name in enumerate(comparison_result['matching_fields'], start=2):
            ws[f'A{idx}'] = field_name
            ws[f'A{idx}'].border = self.border
        
        self._auto_adjust_columns(ws)
    
    def generate_data_comparison_report(self, comparison_result: Dict[str, Any], table_name: str, output_file: str, source_env_name: str = None, target_env_name: str = None):
        """
        Generate Excel report for data comparison
        
        Args:
            comparison_result: Dictionary containing comparison results
            table_name: Name of the table being compared
            output_file: Output Excel file path
            source_env_name: Name of source environment (optional)
            target_env_name: Name of target environment (optional)
        """
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Create Summary sheet
        self._create_data_summary_sheet(wb, comparison_result, table_name, source_env_name, target_env_name)
        
        # Create Records Only in Source sheet
        if comparison_result['only_in_source']:
            self._create_records_only_in_source_sheet(wb, comparison_result, source_env_name)
        
        # Create Records Only in Target sheet
        if comparison_result['only_in_target']:
            self._create_records_only_in_target_sheet(wb, comparison_result, target_env_name)
        
        # Create Field Mismatches sheet
        if comparison_result['mismatches']:
            self._create_field_mismatches_sheet(wb, comparison_result, source_env_name, target_env_name)
        
        # Create Name Matched Records sheet
        if comparison_result.get('name_matched_records'):
            self._create_name_matched_records_sheet(wb, comparison_result, source_env_name, target_env_name)
        
        # Create Matching Records sheet
        if comparison_result['matching_records']:
            self._create_matching_records_sheet(wb, comparison_result)
        
        # Save workbook
        wb.save(output_file)
        print(f"  Excel report saved to: {output_file}")
    
    def _create_data_summary_sheet(self, wb, comparison_result: Dict[str, Any], table_name: str, source_env_name: str = None, target_env_name: str = None):
        """Create data comparison summary sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        # Determine match mode for title
        match_mode = comparison_result.get('match_mode', 'guid').upper()
        
        # Title
        ws['A1'] = f"Data Comparison Report ({match_mode} Matching)"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:B1')
        
        row = 2
        if match_mode == "GUID":
            note = "Comparison is based on GUID matching. System fields (modifiedon, createdby, ownerid, etc.) are excluded."
        elif match_mode == "NAME":
            note = "Comparison is based on Primary Name matching. System fields and primary keys are excluded."
        else:  # BOTH
            note = "Comparison uses GUID first, then Primary Name for unmatched records. System fields excluded."
        ws[f'A{row}'] = f"Note: {note}"
        ws[f'A{row}'].font = Font(italic=True, size=9)
        ws.merge_cells(f'A{row}:B{row}')
        
        # Metadata
        row = 4
        ws[f'A{row}'] = "Table Name:"
        ws[f'B{row}'] = table_name
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        source_label = source_env_name if source_env_name else "Source Environment"
        ws[f'A{row}'] = f"{source_label}:"
        ws[f'B{row}'] = comparison_result['source_url']
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        target_label = target_env_name if target_env_name else "Target Environment"
        ws[f'A{row}'] = f"{target_label}:"
        ws[f'B{row}'] = comparison_result['target_url']
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Match Mode:"
        ws[f'B{row}'] = match_mode
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True, color="0066CC")
        
        row += 1
        ws[f'A{row}'] = "Report Generated:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws[f'A{row}'].font = Font(bold=True)
        
        # Summary statistics
        row += 2
        ws[f'A{row}'] = "COMPARISON SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Count"
        self._apply_header_style(ws, row, 2)
        
        # Statistics rows - base stats
        source_label = source_env_name if source_env_name else "Source"
        target_label = target_env_name if target_env_name else "Target"
        unique_mismatch_records = len(set([m['record_id'] for m in comparison_result['mismatches']])) if comparison_result['mismatches'] else 0
        
        stats = [
            (f"{source_label} Records (Total)", comparison_result['source_record_count'], None),
            (f"{target_label} Records (Total)", comparison_result['target_record_count'], None),
        ]
        
        # Match mode specific stats
        if match_mode == "GUID":
            stats.extend([
                (f"GUIDs Only in {source_label}", len(comparison_result['only_in_source']), self.warning_fill),
                (f"GUIDs Only in {target_label}", len(comparison_result['only_in_target']), self.warning_fill),
                ("GUIDs with Attribute Mismatches", unique_mismatch_records, self.info_fill),
                ("Total Field Mismatches", len(comparison_result['mismatches']), self.info_fill),
                ("Lookup Field (GUID) Mismatches", len(comparison_result.get('guid_mismatches', [])), self.warning_fill),
                ("Same Name, Different GUIDs", len(comparison_result.get('name_matches_with_different_ids', [])), self.warning_fill),
                ("Matching GUIDs (Identical)", len(comparison_result['matching_records']), self.success_fill)
            ])
        elif match_mode == "NAME":
            stats.extend([
                (f"Names Only in {source_label}", len(comparison_result['only_in_source']), self.warning_fill),
                (f"Names Only in {target_label}", len(comparison_result['only_in_target']), self.warning_fill),
                ("Names with Attribute Mismatches", unique_mismatch_records, self.info_fill),
                ("Total Field Mismatches", len(comparison_result['mismatches']), self.info_fill),
                ("Name Matched (Different GUIDs)", len(comparison_result.get('name_matched_records', [])), self.info_fill),
                ("Duplicate Names in Source", len(comparison_result.get('duplicate_names_source', [])), self.warning_fill),
                ("Duplicate Names in Target", len(comparison_result.get('duplicate_names_target', [])), self.warning_fill),
                ("Matching Names (Identical)", len(comparison_result['matching_records']), self.success_fill)
            ])
        else:  # BOTH
            stats.extend([
                ("GUID Matched Records", len(comparison_result.get('guid_matched_records', [])), self.success_fill),
                ("Name Matched (Different GUIDs)", len(comparison_result.get('name_matched_records', [])), self.info_fill),
                (f"Unmatched in {source_label}", len(comparison_result['only_in_source']), self.warning_fill),
                (f"Unmatched in {target_label}", len(comparison_result['only_in_target']), self.warning_fill),
                ("Records with Mismatches", unique_mismatch_records, self.info_fill),
                ("Total Field Mismatches", len(comparison_result['mismatches']), self.info_fill),
                ("Total Matching (Identical)", len(comparison_result['matching_records']), self.success_fill)
            ])
        
        for label, value, fill in stats:
            row += 1
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            if fill:
                ws[f'B{row}'].fill = fill
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
        
        self._auto_adjust_columns(ws)
    
    def _create_records_only_in_source_sheet(self, wb, comparison_result: Dict[str, Any], source_env_name: str = None):
        """Create sheet for records only in source (by primary name)"""
        source_label = source_env_name if source_env_name else "Source"
        ws = wb.create_sheet(f"Only in {source_label}")
        
        if not comparison_result['only_in_source']:
            return
        
        # Add title
        ws['A1'] = f"Records Present in {source_label} but NOT in Target (Primary Name Matching)"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:C1')
        
        # Get all field names
        first_record = comparison_result['only_in_source'][0]
        fields = [f for f in first_record.keys() if not f.startswith("@")]
        
        # Ensure primary name is first column
        primary_name_field = comparison_result.get('primary_name_field', 'name')
        if primary_name_field in fields:
            fields.remove(primary_name_field)
            fields.insert(0, primary_name_field)
        
        # Headers (starting at row 3)
        for col, field in enumerate(fields, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = field
        
        self._apply_header_style(ws, 3, len(fields))
        
        # Data (starting at row 4)
        for row_idx, record in enumerate(comparison_result['only_in_source'], start=4):
            for col_idx, field in enumerate(fields, start=1):
                value = record.get(field, "")
                # Truncate long values
                if isinstance(value, str) and len(value) > 500:
                    value = value[:500] + "..."
                ws.cell(row=row_idx, column=col_idx).value = str(value) if value else ""
                ws.cell(row=row_idx, column=col_idx).border = self.border
                # Highlight primary name column
                if field == primary_name_field:
                    ws.cell(row=row_idx, column=col_idx).fill = self.info_fill
        
        self._auto_adjust_columns(ws)
    
    def _create_records_only_in_target_sheet(self, wb, comparison_result: Dict[str, Any], target_env_name: str = None):
        """Create sheet for records only in target (by primary name)"""
        target_label = target_env_name if target_env_name else "Target"
        ws = wb.create_sheet(f"Only in {target_label}")
        
        if not comparison_result['only_in_target']:
            return
        
        # Add title
        ws['A1'] = f"Records Present in {target_label} but NOT in Source (Primary Name Matching)"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:C1')
        
        # Get all field names
        first_record = comparison_result['only_in_target'][0]
        fields = [f for f in first_record.keys() if not f.startswith("@")]
        
        # Ensure primary name is first column
        primary_name_field = comparison_result.get('primary_name_field', 'name')
        if primary_name_field in fields:
            fields.remove(primary_name_field)
            fields.insert(0, primary_name_field)
        
        # Headers (starting at row 3)
        for col, field in enumerate(fields, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = field
        
        self._apply_header_style(ws, 3, len(fields))
        
        # Data (starting at row 4)
        for row_idx, record in enumerate(comparison_result['only_in_target'], start=4):
            for col_idx, field in enumerate(fields, start=1):
                value = record.get(field, "")
                # Truncate long values
                if isinstance(value, str) and len(value) > 500:
                    value = value[:500] + "..."
                ws.cell(row=row_idx, column=col_idx).value = str(value) if value else ""
                ws.cell(row=row_idx, column=col_idx).border = self.border
                # Highlight primary name column
                if field == primary_name_field:
                    ws.cell(row=row_idx, column=col_idx).fill = self.info_fill
        
        self._auto_adjust_columns(ws)
    
    def _create_field_mismatches_sheet(self, wb, comparison_result: Dict[str, Any], source_env_name: str = None, target_env_name: str = None):
        """Create sheet for field mismatches (name-based matching)"""
        ws = wb.create_sheet("Attribute Mismatches")

        # Add title
        ws['A1'] = "Records Matched by Primary Name with Attribute Differences"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:F1')
        
        row = 2
        ws[f'A{row}'] = "Note: System fields are excluded from comparison. Primary keys (GUIDs) are not used for matching."
        ws[f'A{row}'].font = Font(italic=True, size=9)
        ws.merge_cells(f'A{row}:F{row}')
        
        # Headers (starting at row 4)
        source_label = source_env_name if source_env_name else "Source"
        target_label = target_env_name if target_env_name else "Target"
        row = 4
        headers = ["Source ID", "Primary Name", "Field Name", f"{source_label} Value", f"{target_label} Value", "Type"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
        
        self._apply_header_style(ws, row, len(headers))
        
        # Data (starting at row 5)
        row = 5
        for mismatch in comparison_result['mismatches']:
            ws[f'A{row}'] = str(mismatch['record_id'])
            ws[f'B{row}'] = str(mismatch.get('record_name', ''))
            ws[f'C{row}'] = mismatch['field_name']
            
            source_val = mismatch['source_value']
            target_val = mismatch['target_value']
            
            # Truncate long values
            if isinstance(source_val, str) and len(source_val) > 500:
                source_val = source_val[:500] + "..."
            if isinstance(target_val, str) and len(target_val) > 500:
                target_val = target_val[:500] + "..."
            
            ws[f'D{row}'] = str(source_val) if source_val else ""
            ws[f'E{row}'] = str(target_val) if target_val else ""
            
            # Mark GUID fields
            is_guid = mismatch.get('is_guid', False)
            ws[f'F{row}'] = "GUID/Lookup" if is_guid else "Regular"
            
            # Apply borders and highlighting
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
            
            # Highlight GUIDs differently
            if is_guid:
                ws[f'D{row}'].fill = self.warning_fill
                ws[f'E{row}'].fill = self.warning_fill
                ws[f'F{row}'].fill = self.warning_fill
            else:
                ws[f'D{row}'].fill = self.info_fill
                ws[f'E{row}'].fill = self.info_fill
            
            row += 1
        
        self._auto_adjust_columns(ws)
    
    def _create_matching_records_sheet(self, wb, comparison_result: Dict[str, Any]):
        """Create sheet for matching records (name-based)"""
        ws = wb.create_sheet("Matching Records")

        # Add title
        ws['A1'] = "Records Matched by Primary Name with Identical Attribute Values"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:B1')

        # Headers (starting at row 3)
        ws['A3'] = "Source ID"
        ws['A3'].fill = self.success_fill
        ws['A3'].font = Font(bold=True)
        ws['A3'].border = self.border
        
        # Data (starting at row 4)
        for idx, record_id in enumerate(comparison_result['matching_records'], start=4):
            ws[f'A{idx}'] = str(record_id)
            ws[f'A{idx}'].border = self.border
            ws[f'A{idx}'].fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        
        self._auto_adjust_columns(ws)
    
    def _create_guid_mismatches_sheet(self, wb, comparison_result: Dict[str, Any], source_env_name: str = None, target_env_name: str = None):
        """Create sheet specifically for GUID/Lookup field mismatches"""
        ws = wb.create_sheet("GUID Mismatches")
        
        # Title
        ws['A1'] = "GUID/Lookup Field Mismatches"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:E1')
        
        source_label = source_env_name if source_env_name else "Source"
        target_label = target_env_name if target_env_name else "Target"
        
        row = 3
        ws[f'A{row}'] = f"These are lookup/relationship fields that point to different records in {source_label} vs {target_label}."
        ws[f'A{row}'].font = Font(italic=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        # Headers
        row += 2
        headers = ["Record ID", "Record Name", "Field Name", f"{source_label} GUID", f"{target_label} GUID"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
        
        self._apply_header_style(ws, row, len(headers))
        
        # Data
        for mismatch in comparison_result['guid_mismatches']:
            row += 1
            ws[f'A{row}'] = str(mismatch['record_id'])
            ws[f'B{row}'] = str(mismatch.get('record_name', ''))
            ws[f'C{row}'] = mismatch['field_name']
            ws[f'D{row}'] = str(mismatch['source_value']) if mismatch['source_value'] else ""
            ws[f'E{row}'] = str(mismatch['target_value']) if mismatch['target_value'] else ""
            
            # Apply borders and highlighting
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
            
            ws[f'D{row}'].fill = self.warning_fill
            ws[f'E{row}'].fill = self.warning_fill
        
        self._auto_adjust_columns(ws)
    
    def _create_name_id_mismatch_sheet(self, wb, comparison_result: Dict[str, Any], source_env_name: str = None, target_env_name: str = None):
        """Create sheet for records with same name but different IDs"""
        ws = wb.create_sheet("Name-ID Conflicts")
        
        # Title
        ws['A1'] = "Records with Same Name but Different IDs"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:D1')
        
        source_label = source_env_name if source_env_name else "Source"
        target_label = target_env_name if target_env_name else "Target"
        
        row = 3
        ws[f'A{row}'] = "These records have the same primary name field but different IDs. This may indicate duplicates or migration issues."
        ws[f'A{row}'].font = Font(italic=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        # Headers
        row += 2
        headers = [f"{source_label} ID", f"{target_label} ID", f"{comparison_result.get('primary_name_field', 'Name')}", "Status"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
        
        self._apply_header_style(ws, row, len(headers))
        
        # Data
        for match in comparison_result['name_matches_with_different_ids']:
            row += 1
            ws[f'A{row}'] = str(match['source_id'])
            ws[f'B{row}'] = str(match['target_id'])
            ws[f'C{row}'] = str(match['name'])
            ws[f'D{row}'] = match['status']
            
            # Apply borders and highlighting
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                cell.fill = self.warning_fill
        
        self._auto_adjust_columns(ws)
    
    def _create_name_matched_records_sheet(self, wb, comparison_result: Dict[str, Any], source_env_name: str = None, target_env_name: str = None):
        """Create sheet for records matched by name (with different GUIDs)"""
        ws = wb.create_sheet("Name Matched Records")
        
        # Title
        ws['A1'] = "Records Matched by Primary Name (Different GUIDs)"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:E1')
        
        source_label = source_env_name if source_env_name else "Source"
        target_label = target_env_name if target_env_name else "Target"
        
        row = 3
        ws[f'A{row}'] = "These records were matched using their primary name field because GUIDs differ between environments."
        ws[f'A{row}'].font = Font(italic=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws[f'A{row}'] = "This is common in independent environments or after data migrations where GUIDs are regenerated."
        ws[f'A{row}'].font = Font(italic=True, size=9)
        ws.merge_cells(f'A{row}:E{row}')
        
        # Headers
        row += 2
        primary_name_field = comparison_result.get('primary_name_field', 'Name')
        headers = [primary_name_field, f"{source_label} GUID", f"{target_label} GUID", "Has Differences", "Status"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
        
        self._apply_header_style(ws, row, len(headers))
        
        # Data
        name_matched = comparison_result.get('name_matched_records', [])
        for match in name_matched:
            row += 1
            ws[f'A{row}'] = str(match.get('name', ''))
            ws[f'B{row}'] = str(match.get('source_id', ''))
            ws[f'C{row}'] = str(match.get('target_id', ''))
            has_diff = match.get('has_differences', False)
            ws[f'D{row}'] = "Yes" if has_diff else "No"
            ws[f'E{row}'] = "Matched by Name" if not has_diff else "Matched by Name (with attribute differences)"
            
            # Apply borders and highlighting
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
            
            # Highlight based on differences
            if has_diff:
                ws[f'D{row}'].fill = self.info_fill
                ws[f'E{row}'].fill = self.info_fill
            else:
                ws[f'D{row}'].fill = self.success_fill
                ws[f'E{row}'].fill = self.success_fill
        
        # Add summary at bottom
        row += 2
        total_name_matched = len(name_matched)
        with_diff = len([m for m in name_matched if m.get('has_differences')])
        without_diff = total_name_matched - with_diff
        
        ws[f'A{row}'] = "Summary:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = f"Total Name Matched: {total_name_matched}"
        row += 1
        ws[f'A{row}'] = f"Identical (no differences): {without_diff}"
        row += 1
        ws[f'A{row}'] = f"With attribute differences: {with_diff}"
        
        self._auto_adjust_columns(ws)
    
    # NOTE: Related entity (child) comparisons have been removed in simplified mode
    
    def generate_flow_comparison_report(self, comparison_result: Dict[str, Any], output_file: str, source_env_name: str = None, target_env_name: str = None):
        """
        Generate Excel report for flow comparison
        
        Args:
            comparison_result: Dictionary containing comparison results
            output_file: Output Excel file path
            source_env_name: Name of source environment (optional)
            target_env_name: Name of target environment (optional)
        """
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Create Summary sheet
        self._create_flow_summary_sheet(wb, comparison_result, source_env_name, target_env_name)
        
        # Create Identical Flows sheet
        if comparison_result.get('identical_flows'):
            self._create_identical_flows_sheet(wb, comparison_result)
        
        # Create Different Flows sheet
        if comparison_result.get('non_identical_flows'):
            self._create_different_flows_sheet(wb, comparison_result, source_env_name, target_env_name)
        
        # Create Missing in Target sheet
        if comparison_result.get('missing_in_target'):
            self._create_missing_flows_sheet(wb, comparison_result, target_env_name)
        
        # Create Errors sheet
        if comparison_result.get('errors'):
            self._create_flow_errors_sheet(wb, comparison_result)
        
        # Save workbook
        wb.save(output_file)
        print(f"  Excel report saved to: {output_file}")
    
    def _create_flow_summary_sheet(self, wb, comparison_result: Dict[str, Any], source_env_name: str = None, target_env_name: str = None):
        """Create flow comparison summary sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "Flow Comparison Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:B1')
        
        # Metadata
        row = 3
        source_label = source_env_name if source_env_name else "Source Environment"
        ws[f'A{row}'] = f"{source_label}:"
        ws[f'B{row}'] = comparison_result['source_url']
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        target_label = target_env_name if target_env_name else "Target Environment"
        ws[f'A{row}'] = f"{target_label}:"
        ws[f'B{row}'] = comparison_result['target_url']
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Report Generated:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws[f'A{row}'].font = Font(bold=True)
        
        # Summary statistics
        row += 2
        ws[f'A{row}'] = "COMPARISON SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Count"
        self._apply_header_style(ws, row, 2)
        
        # Statistics rows
        source_label = source_env_name if source_env_name else "Source"
        target_label = target_env_name if target_env_name else "Target"
        stats = [
            (f"{source_label} Flows (Total)", comparison_result['source_count'], None),
            (f"{target_label} Flows (Total)", comparison_result['target_count'], None),
            ("Identical Flows", len(comparison_result.get('identical_flows', [])), self.success_fill),
            ("Different Flows", len(comparison_result.get('non_identical_flows', [])), self.warning_fill),
            (f"Missing in {target_label}", comparison_result['missing_in_target_count'], self.warning_fill),
            ("Flows with Errors", comparison_result['error_count'], self.warning_fill)
        ]
        
        for label, value, fill in stats:
            row += 1
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            if fill:
                ws[f'B{row}'].fill = fill
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
        
        self._auto_adjust_columns(ws)
    
    def _create_identical_flows_sheet(self, wb, comparison_result: Dict[str, Any]):
        """Create sheet for identical flows"""
        ws = wb.create_sheet("Identical Flows")
        
        # Title
        ws['A1'] = "Flows with Identical Definitions"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:B1')
        
        # Headers
        ws['A3'] = "Flow Name"
        ws['A3'].fill = self.success_fill
        ws['A3'].font = Font(bold=True)
        ws['A3'].border = self.border
        
        # Data
        for idx, flow_name in enumerate(comparison_result['identical_flows'], start=4):
            ws[f'A{idx}'] = flow_name
            ws[f'A{idx}'].border = self.border
            ws[f'A{idx}'].fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        
        self._auto_adjust_columns(ws)
    
    def _create_different_flows_sheet(self, wb, comparison_result: Dict[str, Any], source_env_name: str = None, target_env_name: str = None):
        """Create sheet for flows with differences"""
        ws = wb.create_sheet("Different Flows")
        
        # Title
        ws['A1'] = "Flows with Different Definitions"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:F1')
        
        source_label = source_env_name if source_env_name else "Source"
        target_label = target_env_name if target_env_name else "Target"
        
        # Headers
        row = 3
        headers = ["Flow Name", f"{source_label} Hash", f"{target_label} Hash", "Added Paths", "Removed Paths", "Changed Paths"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
        
        self._apply_header_style(ws, row, len(headers))
        
        # Data
        for comparison in comparison_result['comparisons']:
            if comparison['status'] != 'different':
                continue
            
            row += 1
            ws[f'A{row}'] = comparison['name']
            ws[f'B{row}'] = comparison['source']['hash'][:16] + "..." if comparison['source']['hash'] else ""
            ws[f'C{row}'] = comparison['target']['hash'][:16] + "..." if comparison['target']['hash'] else ""
            
            # Diff details
            if comparison.get('diff'):
                ws[f'D{row}'] = len(comparison['diff']['added'])
                ws[f'E{row}'] = len(comparison['diff']['removed'])
                ws[f'F{row}'] = len(comparison['diff']['changed'])
            else:
                ws[f'D{row}'] = "N/A"
                ws[f'E{row}'] = "N/A"
                ws[f'F{row}'] = "N/A"
            
            # Apply borders and highlighting
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
            
            ws[f'D{row}'].fill = self.info_fill
            ws[f'E{row}'].fill = self.info_fill
            ws[f'F{row}'].fill = self.warning_fill
        
        self._auto_adjust_columns(ws)
        
        # Create detailed action differences sheet
        self._create_action_differences_sheet(wb, comparison_result, source_env_name, target_env_name)
    
    def _create_action_differences_sheet(self, wb, comparison_result: Dict[str, Any], source_env_name: str = None, target_env_name: str = None):
        """Create sheet for action-level differences"""
        ws = wb.create_sheet("Action Differences")
        
        # Title
        ws['A1'] = "Flow Action-Level Differences"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:E1')
        
        source_label = source_env_name if source_env_name else "Source"
        target_label = target_env_name if target_env_name else "Target"
        
        # Headers
        row = 3
        headers = ["Flow Name", "Action Name", "Status", "Property Path", f"{source_label} Value", f"{target_label} Value"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
        
        self._apply_header_style(ws, row, len(headers))
        
        # Data
        for comparison in comparison_result['comparisons']:
            if not comparison.get('action_differences'):
                continue
            
            for action_diff in comparison['action_differences']:
                action_name = action_diff['action_name']
                status = action_diff['status']
                
                if status in ['added', 'removed']:
                    # Simple row for added/removed actions
                    row += 1
                    ws[f'A{row}'] = comparison['name']
                    ws[f'B{row}'] = action_name
                    ws[f'C{row}'] = status.upper()
                    ws[f'D{row}'] = ""
                    ws[f'E{row}'] = "(removed)" if status == 'removed' else ""
                    ws[f'F{row}'] = "(added)" if status == 'added' else ""
                    
                    for col in range(1, 7):
                        cell = ws.cell(row=row, column=col)
                        cell.border = self.border
                        if col in [2, 3]:
                            cell.fill = self.warning_fill
                else:
                    # Changed action - show each property change
                    for prop_change in action_diff['changed_properties']:
                        row += 1
                        ws[f'A{row}'] = comparison['name']
                        ws[f'B{row}'] = action_name
                        ws[f'C{row}'] = "CHANGED"
                        ws[f'D{row}'] = prop_change['path']
                        
                        source_val = prop_change['source_value']
                        target_val = prop_change['target_value']
                        
                        # Truncate long values
                        if isinstance(source_val, str) and len(source_val) > 200:
                            source_val = source_val[:200] + "..."
                        if isinstance(target_val, str) and len(target_val) > 200:
                            target_val = target_val[:200] + "..."
                        
                        ws[f'E{row}'] = str(source_val)
                        ws[f'F{row}'] = str(target_val)
                        
                        for col in range(1, 7):
                            cell = ws.cell(row=row, column=col)
                            cell.border = self.border
                        
                        ws[f'E{row}'].fill = self.info_fill
                        ws[f'F{row}'].fill = self.info_fill
        
        self._auto_adjust_columns(ws)
    
    def _create_missing_flows_sheet(self, wb, comparison_result: Dict[str, Any], target_env_name: str = None):
        """Create sheet for flows missing in target"""
        ws = wb.create_sheet("Missing in Target")
        
        target_label = target_env_name if target_env_name else "Target"
        
        # Title
        ws['A1'] = f"Flows Missing in {target_label}"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:C1')
        
        # Headers
        ws['A3'] = "Flow Name"
        ws['B3'] = "Flow ID"
        ws['C3'] = "Status"
        self._apply_header_style(ws, 3, 3)
        
        # Data
        row = 4
        for comparison in comparison_result['comparisons']:
            if comparison['status'] == 'missing_in_target':
                ws[f'A{row}'] = comparison['name']
                ws[f'B{row}'] = comparison['source']['flow_id']
                ws[f'C{row}'] = "Missing in Target"
                
                for col in range(1, 4):
                    cell = ws.cell(row=row, column=col)
                    cell.border = self.border
                    cell.fill = self.warning_fill
                
                row += 1
        
        self._auto_adjust_columns(ws)
    
    def _create_flow_errors_sheet(self, wb, comparison_result: Dict[str, Any]):
        """Create sheet for flows with errors"""
        ws = wb.create_sheet("Errors")
        
        # Title
        ws['A1'] = "Flows with Processing Errors"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:D1')
        
        # Headers
        ws['A3'] = "Flow Name"
        ws['B3'] = "Source Error"
        ws['C3'] = "Target Error"
        ws['D3'] = "Status"
        self._apply_header_style(ws, 3, 4)
        
        # Data
        row = 4
        for comparison in comparison_result['comparisons']:
            if comparison['status'] == 'error':
                ws[f'A{row}'] = comparison['name']
                ws[f'B{row}'] = comparison['source'].get('error', '') if comparison['source'] else ''
                ws[f'C{row}'] = comparison['target'].get('error', '') if comparison['target'] else ''
                ws[f'D{row}'] = "Error"
                
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.border = self.border
                    cell.fill = self.warning_fill
                
                row += 1
        
        self._auto_adjust_columns(ws)
    
    def generate_solution_comparison_report(self, comparison_result: Dict[str, Any], output_file: str, source_env_name: str = None, target_env_name: str = None):
        """
        Generate Excel report for solution comparison
        
        Args:
            comparison_result: Dictionary containing comparison results
            output_file: Output Excel file path
            source_env_name: Name of source environment (optional)
            target_env_name: Name of target environment (optional)
        """
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Create Summary sheet
        self._create_solution_summary_sheet(wb, comparison_result, source_env_name, target_env_name)
        
        # Create Component Type Summary sheet
        self._create_component_type_summary_sheet(wb, comparison_result, source_env_name, target_env_name)
        
        # Create Components Only in Source sheet
        if comparison_result.get('only_in_source'):
            self._create_components_only_in_source_sheet(wb, comparison_result, source_env_name)
        
        # Create Components Only in Target sheet
        if comparison_result.get('only_in_target'):
            self._create_components_only_in_target_sheet(wb, comparison_result, target_env_name)
        
        # Create Common Components sheet
        if comparison_result.get('common_components'):
            self._create_common_components_sheet(wb, comparison_result)
        
        # Save workbook
        wb.save(output_file)
        print(f"  Excel report saved to: {output_file}")
    
    def _create_solution_summary_sheet(self, wb, comparison_result: Dict[str, Any], source_env_name: str = None, target_env_name: str = None):
        """Create solution comparison summary sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "Solution Comparison Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:B1')
        
        # Metadata
        row = 3
        ws[f'A{row}'] = "Solution Name:"
        ws[f'B{row}'] = comparison_result['solution_name']
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        source_label = source_env_name if source_env_name else "Source Environment"
        ws[f'A{row}'] = f"{source_label}:"
        ws[f'B{row}'] = comparison_result['source_url']
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        target_label = target_env_name if target_env_name else "Target Environment"
        ws[f'A{row}'] = f"{target_label}:"
        ws[f'B{row}'] = comparison_result['target_url']
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Report Generated:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws[f'A{row}'].font = Font(bold=True)
        
        # Solution details
        if comparison_result.get('source_solution'):
            row += 2
            ws[f'A{row}'] = f"{source_label} Solution Details"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:B{row}')
            
            source_sol = comparison_result['source_solution']
            row += 1
            ws[f'A{row}'] = "Version:"
            ws[f'B{row}'] = source_sol.get('version', 'N/A')
            
            row += 1
            ws[f'A{row}'] = "Is Managed:"
            ws[f'B{row}'] = "Yes" if source_sol.get('ismanaged') else "No"
            
            row += 1
            ws[f'A{row}'] = "Friendly Name:"
            ws[f'B{row}'] = source_sol.get('friendlyname', 'N/A')
        
        if comparison_result.get('target_solution'):
            row += 2
            ws[f'A{row}'] = f"{target_label} Solution Details"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:B{row}')
            
            target_sol = comparison_result['target_solution']
            row += 1
            ws[f'A{row}'] = "Version:"
            ws[f'B{row}'] = target_sol.get('version', 'N/A')
            
            row += 1
            ws[f'A{row}'] = "Is Managed:"
            ws[f'B{row}'] = "Yes" if target_sol.get('ismanaged') else "No"
            
            row += 1
            ws[f'A{row}'] = "Friendly Name:"
            ws[f'B{row}'] = target_sol.get('friendlyname', 'N/A')
        
        # Summary statistics
        row += 2
        ws[f'A{row}'] = "COMPARISON SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Count"
        self._apply_header_style(ws, row, 2)
        
        # Statistics rows
        source_label = source_env_name if source_env_name else "Source"
        target_label = target_env_name if target_env_name else "Target"
        stats = [
            (f"{source_label} Components (Total)", comparison_result['source_component_count'], None),
            (f"{target_label} Components (Total)", comparison_result['target_component_count'], None),
            (f"Components Only in {source_label}", comparison_result['source_only_count'], self.warning_fill),
            (f"Components Only in {target_label}", comparison_result['target_only_count'], self.warning_fill),
            ("Common Components", comparison_result['common_count'], self.success_fill)
        ]
        
        for label, value, fill in stats:
            row += 1
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            if fill:
                ws[f'B{row}'].fill = fill
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
        
        self._auto_adjust_columns(ws)
    
    def _create_component_type_summary_sheet(self, wb, comparison_result: Dict[str, Any], source_env_name: str = None, target_env_name: str = None):
        """Create component type summary sheet"""
        ws = wb.create_sheet("Component Type Summary")
        
        # Title
        ws['A1'] = "Solution Components by Type"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:F1')
        
        source_label = source_env_name if source_env_name else "Source"
        target_label = target_env_name if target_env_name else "Target"
        
        # Headers
        row = 3
        headers = ["Component Type", f"{source_label} Count", f"{target_label} Count", "Common", f"Only in {source_label}", f"Only in {target_label}"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
        
        self._apply_header_style(ws, row, len(headers))
        
        # Data - sort by component type name
        component_summary = comparison_result.get('component_summary', {})
        for comp_type in sorted(component_summary.keys()):
            stats = component_summary[comp_type]
            row += 1
            ws[f'A{row}'] = comp_type
            ws[f'B{row}'] = stats['source']
            ws[f'C{row}'] = stats['target']
            ws[f'D{row}'] = stats['common']
            ws[f'E{row}'] = stats['source_only']
            ws[f'F{row}'] = stats['target_only']
            
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
            
            # Highlight differences
            if stats['source_only'] > 0:
                ws[f'E{row}'].fill = self.warning_fill
            if stats['target_only'] > 0:
                ws[f'F{row}'].fill = self.warning_fill
            if stats['common'] > 0:
                ws[f'D{row}'].fill = self.success_fill
        
        self._auto_adjust_columns(ws)
    
    def _create_components_only_in_source_sheet(self, wb, comparison_result: Dict[str, Any], source_env_name: str = None):
        """Create sheet for components only in source"""
        source_label = source_env_name if source_env_name else "Source"
        ws = wb.create_sheet(f"Only in {source_label}")
        
        # Title
        ws['A1'] = f"Components Only in {source_label}"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:C1')
        
        # Headers
        ws['A3'] = "Component Type"
        ws['B3'] = "Object ID"
        ws['C3'] = "Root Behavior"
        self._apply_header_style(ws, 3, 3)
        
        # Data - sort by component type
        components = sorted(comparison_result['only_in_source'], key=lambda x: x['componenttype_name'])
        
        row = 4
        for comp in components:
            ws[f'A{row}'] = comp['componenttype_name']
            ws[f'B{row}'] = comp['objectid']
            ws[f'C{row}'] = str(comp.get('rootcomponentbehavior', 'N/A'))
            
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                cell.fill = self.warning_fill
            
            row += 1
        
        self._auto_adjust_columns(ws)
    
    def _create_components_only_in_target_sheet(self, wb, comparison_result: Dict[str, Any], target_env_name: str = None):
        """Create sheet for components only in target"""
        target_label = target_env_name if target_env_name else "Target"
        ws = wb.create_sheet(f"Only in {target_label}")
        
        # Title
        ws['A1'] = f"Components Only in {target_label}"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:C1')
        
        # Headers
        ws['A3'] = "Component Type"
        ws['B3'] = "Object ID"
        ws['C3'] = "Root Behavior"
        self._apply_header_style(ws, 3, 3)
        
        # Data - sort by component type
        components = sorted(comparison_result['only_in_target'], key=lambda x: x['componenttype_name'])
        
        row = 4
        for comp in components:
            ws[f'A{row}'] = comp['componenttype_name']
            ws[f'B{row}'] = comp['objectid']
            ws[f'C{row}'] = str(comp.get('rootcomponentbehavior', 'N/A'))
            
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                cell.fill = self.info_fill
            
            row += 1
        
        self._auto_adjust_columns(ws)
    
    def _create_common_components_sheet(self, wb, comparison_result: Dict[str, Any]):
        """Create sheet for common components"""
        ws = wb.create_sheet("Common Components")
        
        # Title
        ws['A1'] = "Components Present in Both Environments"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:C1')
        
        # Headers
        ws['A3'] = "Component Type"
        ws['B3'] = "Object ID"
        ws['C3'] = "Root Behavior"
        self._apply_header_style(ws, 3, 3)
        
        # Data - sort by component type
        components = sorted(comparison_result['common_components'], key=lambda x: x['componenttype_name'])
        
        row = 4
        for comp in components:
            ws[f'A{row}'] = comp['componenttype_name']
            ws[f'B{row}'] = comp['objectid']
            ws[f'C{row}'] = str(comp.get('rootcomponentbehavior', 'N/A'))
            
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            
            row += 1
        
        self._auto_adjust_columns(ws)
